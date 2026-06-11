"""
Full CTE import from Excel — My Business + Lead Gen
Run: source venv/bin/activate && python3 import_cte.py
"""
import sys, os, re
from datetime import datetime, date

sys.path.insert(0, os.path.dirname(__file__))
XLS = '/Users/edentdg/.hermes/cache/documents/doc_d85cd1f1d70e_CTE_2026_Year_2.xlsx'

import openpyxl
from app import create_app, db
from app.models import Transaction, Agent, LeadGenLog

app = create_app()

# ── Helpers ──────────────────────────────────────────────────────────

def cm(val):
    """Clean money → float or None"""
    if val is None: return None
    s = re.sub(r'[\$,\s"]', '', str(val))
    try: return float(s) if s else None
    except: return None

def cp(val):
    """Clean percent → float or None"""
    if val is None: return None
    s = str(val).replace('%','').strip()
    try: return float(s) if s else None
    except: return None

def cd(val):
    """Clean date → date object or None"""
    if val is None: return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    if s in ('', 'FALSE', '#VALUE!', '#REF!', '#N/A'): return None
    for fmt in ('%m/%d/%Y','%m/%d/%y','%-m/%-d/%Y','%-m/%-d/%y'):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    return None

def cs(val, maxlen=None):
    """Clean string"""
    if val is None: return None
    s = str(val).strip().strip('"')
    if s in ('FALSE','TRUE','#VALUE!','#REF!','#N/A','0',''): return None
    if maxlen: s = s[:maxlen]
    return s or None

STATUS_MAP = {
    '`pending':        'Pending',
    'pending':         'Pending',
    'signed':          'Pre-Signed',
    'loi':             'Pre-Signed',
    'unsigned':        'Pre-Signed',
    'temp off mark':   'Temp Off Market',
    'temp off market': 'Temp Off Market',
    'y-sale failed':   'y-Sale Failed',
    'x-cancelled':     'x-Cancelled',
    'z-expired':       'z-Expired',
    'closed':          'Closed',
    'active':          'Active',
    'pipeline':        'Pipeline',
    'coming soon':     'Coming Soon',
}
VALID_STATUSES = set(STATUS_MAP.values()) | {'Active','Pending','Closed','Pipeline',
    'Pre-Signed','Coming Soon','x-Cancelled','y-Sale Failed','z-Expired','Temp Off Market'}

def norm_status(raw):
    if not raw: return None
    k = str(raw).strip().lower()
    return STATUS_MAP.get(k, str(raw).strip())

# ── My Business Import ───────────────────────────────────────────────
# Column index map — update here if CTE spreadsheet columns change
C_LIST_FLAG    = 2;  C_BUYER_FLAG   = 3
C_STATUS       = 23; C_SUB_STATUS   = 24; C_ADDRESS      = 25
C_CLIENT       = 26; C_SOURCE       = 27; C_SIGNED       = 28
C_MLS_LIVE     = 29; C_EXP_DATE     = 30; C_UNDER_CONTR  = 31
C_PROJ_CLOSE   = 32; C_CLOSE_DATE   = 33; C_LIST_PRICE   = 34
C_SALE_PRICE   = 35; C_COMM_PCT     = 37; C_GCI          = 39
C_BONUS        = 40; C_TX_FEE       = 41; C_BROKER_SPLIT = 42
C_FRANCHISE    = 43; C_REFERRAL     = 44; C_PRI_AGENT    = 45
C_PRI_PCT      = 46; C_PRIMARY_GCI  = 47; C_SEC_AGENT    = 48
C_SEC_PCT      = 49; C_SEC_GCI      = 50; C_MORTGAGE     = 64
C_TITLE        = 65; C_LOCATION     = 68; C_NET_INCOME   = 72
C_TAXES        = 73; C_NET_AFTER_TX = 74

def import_my_business(wb, agents_by_name):
    ws = wb['My Business']
    rows = list(ws.iter_rows(min_row=22, values_only=True))  # data starts row 22

    imported = skipped = 0
    errors = []

    for i, row in enumerate(rows):
        if not row or len(row) < 26: skipped += 1; continue

        address = cs(row[C_ADDRESS], 300)
        status_raw = cs(row[C_STATUS])
        status = norm_status(status_raw)

        if not address or not status or status not in VALID_STATUSES:
            skipped += 1; continue

        # Skip dupes
        if Transaction.query.filter_by(address=address, status=status).first():
            skipped += 1; continue

        # Type — use sub_status col and list/buyer flag cols
        sub_status = cs(row[C_SUB_STATUS])
        list_flag  = str(row[C_LIST_FLAG]).strip().lower() if row[C_LIST_FLAG] is not None else ''
        buyer_flag = str(row[C_BUYER_FLAG]).strip().lower() if row[C_BUYER_FLAG] is not None else ''
        if sub_status and 'cre' in sub_status.lower():
            tx_type = 'Commercial'
        elif buyer_flag == 'false' and list_flag != 'false':
            tx_type = 'Listing'
        elif list_flag == 'false' and buyer_flag != 'false':
            tx_type = 'Buyer'
        elif sub_status and 'buyer' in sub_status.lower():
            tx_type = 'Buyer'
        elif sub_status and 'listing' in sub_status.lower():
            tx_type = 'Listing'
        else:
            tx_type = 'Other'

        pri_name = cs(row[C_PRI_AGENT], 100)
        pri_id   = agents_by_name.get(pri_name.strip().lower()) if pri_name else None
        sec_name = cs(row[C_SEC_AGENT], 100)

        close_dt = cd(row[C_CLOSE_DATE])
        year  = close_dt.year  if close_dt else 2026
        month = close_dt.month if close_dt else None

        try:
            tx = Transaction(
                status               = status,
                sub_status           = sub_status,
                transaction_type     = tx_type,
                address              = address,
                client_name          = cs(row[C_CLIENT], 200),
                lead_source          = cs(row[C_SOURCE], 100),
                signed_date          = cd(row[C_SIGNED]),
                mls_live_date        = cd(row[C_MLS_LIVE]),
                expiry_date          = cd(row[C_EXP_DATE]),
                under_contract_date  = cd(row[C_UNDER_CONTR]),
                projected_close_date = cd(row[C_PROJ_CLOSE]),
                close_date           = close_dt,
                list_price           = cm(row[C_LIST_PRICE]),
                sale_price           = cm(row[C_SALE_PRICE]),
                commission_pct       = cp(row[C_COMM_PCT]),
                gci                  = cm(row[C_GCI]) or 0.0,
                bonus                = cm(row[C_BONUS]),
                transaction_fee      = cm(row[C_TX_FEE]),
                broker_split         = cm(row[C_BROKER_SPLIT]),
                franchise_split      = cm(row[C_FRANCHISE]),
                referral_fee         = cm(row[C_REFERRAL]),
                primary_agent_id     = pri_id,
                primary_agent_name   = pri_name,
                primary_agent_pct    = cp(row[C_PRI_PCT]),
                primary_agent_gci    = cm(row[C_PRIMARY_GCI]),
                secondary_agent_name = sec_name,
                secondary_agent_pct  = cp(row[C_SEC_PCT]),
                secondary_agent_gci  = cm(row[C_SEC_GCI]),
                mortgage_company     = cs(row[C_MORTGAGE], 100),
                title_company        = cs(row[C_TITLE], 100),
                location             = cs(row[C_LOCATION], 100),
                net_income           = cm(row[C_NET_INCOME]),
                taxes                = cm(row[C_TAXES]),
                net_after_taxes      = cm(row[C_NET_AFTER_TX]),
                year                 = year,
                month                = month,
            )
            db.session.add(tx)
            imported += 1
            if imported % 50 == 0:
                db.session.commit()
                print(f'  ... {imported} transactions committed')
        except Exception as e:
            errors.append(f'Row {i+22}: {e}')
            db.session.rollback()

    db.session.commit()
    return imported, skipped, errors

# ── Lead Gen Import ──────────────────────────────────────────────────

def import_lead_gen(wb, agents_by_name):
    ws = wb['Lead Gen']
    rows = list(ws.iter_rows(min_row=7, values_only=True))  # data starts row 7

    imported = skipped = 0
    errors = []

    for i, row in enumerate(rows):
        if not row or len(row) < 3: skipped += 1; continue

        log_date = cd(row[1])
        name     = cs(row[2], 100)
        if not log_date or not name or name == 'Name':
            skipped += 1; continue

        agent_id = agents_by_name.get(name.strip().lower())
        if not agent_id:
            # Create agent on the fly if not found
            a = Agent(name=name, status='Active')
            db.session.add(a)
            db.session.flush()
            agent_id = a.id
            agents_by_name[name.strip().lower()] = agent_id

        def ci(val):
            if val is None: return 0
            try: return int(float(str(val)))
            except: return 0

        def cf(val):
            if val is None: return 0.0
            try: return float(str(val))
            except: return 0.0

        try:
            lg = LeadGenLog(
                agent_id           = agent_id,
                log_date           = log_date,
                hours              = cf(row[3]),
                dials              = ci(row[4]),
                contacts           = ci(row[5]),
                nurtures           = ci(row[6]),
                listing_appts_set  = ci(row[7]),
                listing_appts_held = ci(row[8]),
                listings_signed    = ci(row[9]),
                buyer_appts_set    = ci(row[10]),
                buyer_appts_held   = ci(row[11]),
                buyers_signed      = ci(row[12]),
                written_offers     = ci(row[13]),
                showings           = ci(row[14]),
                open_houses        = ci(row[15]),
                lead_source        = cs(row[18], 50),
                notes              = cs(row[19]),
            )
            db.session.add(lg)
            imported += 1
            if imported % 100 == 0:
                db.session.commit()
                print(f'  ... {imported} lead gen rows committed')
        except Exception as e:
            errors.append(f'LG Row {i+7}: {e}')
            db.session.rollback()

    db.session.commit()
    return imported, skipped, errors

# ── Run ──────────────────────────────────────────────────────────────

with app.app_context():
    wb = openpyxl.load_workbook(XLS, read_only=True, data_only=True)

    # Build agent lookup
    agents_by_name = {a.name.strip().lower(): a.id for a in Agent.query.all()}
    print(f'Agents in DB: {len(agents_by_name)}')

    print('\n📋 Importing My Business transactions...')
    t_imp, t_skip, t_err = import_my_business(wb, agents_by_name)
    print(f'  ✅ Imported: {t_imp}  Skipped: {t_skip}  Errors: {len(t_err)}')
    for e in t_err[:5]: print(f'     {e}')

    print('\n📞 Importing Lead Gen logs...')
    lg_imp, lg_skip, lg_err = import_lead_gen(wb, agents_by_name)
    print(f'  ✅ Imported: {lg_imp}  Skipped: {lg_skip}  Errors: {len(lg_err)}')
    for e in lg_err[:5]: print(f'     {e}')

    # Summary counts
    print(f'\n🎉 Done!')
    print(f'   Transactions in DB : {Transaction.query.count()}')
    print(f'   Lead Gen logs in DB: {LeadGenLog.query.count()}')
    print(f'   Agents in DB       : {Agent.query.count()}')
