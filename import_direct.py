"""
Direct psycopg2 import — bypasses Flask/SQLAlchemy entirely.
Run: source venv/bin/activate && python3 import_direct.py
"""
import openpyxl, re, psycopg2
from datetime import datetime, date

XLS = '/Users/edentdg/.hermes/cache/documents/doc_d85cd1f1d70e_CTE_2026_Year_2.xlsx'
PW  = 'SiAmCHSPejkeAVLMAOaZPvUjccxWtTVb'

conn = psycopg2.connect(host='ballast.proxy.rlwy.net', port=34083,
                        dbname='railway', user='postgres', password=PW)
cur = conn.cursor()

# ── Helpers ──────────────────────────────────────────────────────────
def cm(val):
    if val is None: return None
    s = re.sub(r'[\$,\s"]', '', str(val))
    try: return float(s) if s else None
    except: return None

def cp(val):
    if val is None: return None
    s = str(val).replace('%','').strip()
    try: return float(s) if s else None
    except: return None

def cd(val):
    if val is None: return None
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    s = str(val).strip()
    if s in ('','FALSE','#VALUE!','#REF!','#N/A'): return None
    for fmt in ('%m/%d/%Y','%m/%d/%y'):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    return None

def cs(val, maxlen=None):
    if val is None: return None
    s = str(val).strip().strip('"')
    if s in ('FALSE','TRUE','#VALUE!','#REF!','#N/A','0',''): return None
    if maxlen: s = s[:maxlen]
    return s or None

STATUS_MAP = {
    '`pending':'Pending','pending':'Pending','signed':'Pre-Signed',
    'loi':'Pre-Signed','unsigned':'Pre-Signed',
    'temp off mark':'Temp Off Market','temp off market':'Temp Off Market',
    'y-sale failed':'y-Sale Failed','x-cancelled':'x-Cancelled',
    'z-expired':'z-Expired','closed':'Closed','active':'Active','pipeline':'Pipeline',
    'coming soon':'Coming Soon',
}
VALID = {'Active','Pending','Closed','Pipeline','Pre-Signed','Coming Soon',
         'x-Cancelled','y-Sale Failed','z-Expired','Temp Off Market'}

def ns(raw):
    if not raw: return None
    k = str(raw).strip().lower()
    return STATUS_MAP.get(k, str(raw).strip())

# ── Load agents ───────────────────────────────────────────────────────
cur.execute("SELECT id, name FROM agents")
agents = {name.strip().lower(): aid for aid, name in cur.fetchall()}
print(f"Agents in DB: {len(agents)}")

wb = openpyxl.load_workbook(XLS, read_only=True, data_only=True)

# ── MY BUSINESS ───────────────────────────────────────────────────────
# ── MY BUSINESS column index map (CTE 2026 Year.xlsx, row 22+) ───────────
# Update these if the spreadsheet columns ever change — all row[N] refs below use these constants
C_LIST_FLAG   = 2   # Listing indicator
C_BUYER_FLAG  = 3   # Buyer indicator
C_STATUS      = 23  # Status
C_SUB_STATUS  = 24  # Sub Status
C_ADDRESS     = 25  # Property Address
C_CLIENT      = 26  # Client Name
C_SOURCE      = 27  # Lead Source
C_SALE_PRICE  = 28  # Sale Price
C_COMM_PCT    = 29  # Commission %
C_GCI         = 30  # GCI
C_SIGNED      = 31  # Signed Date
C_MLS_LIVE    = 32  # MLS Live Date
C_CLOSE_DATE  = 33  # Close Date
C_PROJ_CLOSE  = 34  # Projected Close
C_EXP_DATE    = 35  # Expiry Date
C_LIST_PRICE  = 36  # List Price
C_BONUS       = 37  # Bonus
C_TX_FEE      = 38  # Transaction Fee
C_BROKER_SPLIT= 39  # Broker Split
C_FRANCHISE   = 40  # Franchise Split
C_REFERRAL    = 41  # Referral Fee
C_PRIMARY_GCI = 47  # Primary Agent GCI (col 48 in 1-indexed)
C_PRI_AGENT   = 45  # Primary Agent Name
C_PRI_PCT     = 46  # Primary Agent %
C_SEC_AGENT   = 48  # Secondary Agent Name
C_SEC_PCT     = 49  # Secondary Agent %
C_SEC_GCI     = 50  # Secondary Agent GCI
C_UNDER_CONTR = 51  # Under Contract Date
C_EO_FEE      = 57  # E&O Fee
C_LEAD_TYPE   = 70  # Lead Type
C_CO_DOLLAR   = 70  # Company Dollar (CTE col 71)
C_1099        = 71  # 1099
# ─────────────────────────────────────────────────────────────────────────────

print('\n📋 Importing My Business...')
ws = wb['My Business']
rows = list(ws.iter_rows(min_row=22, values_only=True))

imported = skipped = 0
for i, row in enumerate(rows):
    if not row or len(row) < 26: skipped += 1; continue
    address    = cs(row[C_ADDRESS], 300)
    status_raw = cs(row[C_STATUS])
    status     = ns(status_raw)
    if not address or not status or status not in VALID: skipped += 1; continue

    # Skip dupes
    cur.execute("SELECT id FROM transactions WHERE address=%s AND status=%s LIMIT 1", (address, status))
    if cur.fetchone(): skipped += 1; continue

    sub_status = cs(row[C_SUB_STATUS])
    list_flag  = str(row[C_LIST_FLAG]).strip().lower() if row[C_LIST_FLAG] is not None else ''
    buyer_flag = str(row[C_BUYER_FLAG]).strip().lower() if row[C_BUYER_FLAG] is not None else ''
    if sub_status and 'cre' in sub_status.lower(): tx_type = 'Commercial'
    elif buyer_flag == 'false' and list_flag != 'false': tx_type = 'Listing'
    elif list_flag == 'false' and buyer_flag != 'false': tx_type = 'Buyer'
    elif sub_status and 'buyer' in sub_status.lower(): tx_type = 'Buyer'
    elif sub_status and 'listing' in sub_status.lower(): tx_type = 'Listing'
    else: tx_type = 'Other'

    pri_name = cs(row[C_PRI_AGENT], 100)
    pri_id   = agents.get(pri_name.strip().lower()) if pri_name else None
    close_dt = cd(row[C_CLOSE_DATE])
    year     = close_dt.year  if close_dt else 2026
    month    = close_dt.month if close_dt else None

    try:
        cur.execute("""
            INSERT INTO transactions (
                status, sub_status, transaction_type, address, client_name, lead_source,
                signed_date, mls_live_date, expiry_date, under_contract_date,
                projected_close_date, close_date, list_price, sale_price,
                commission_pct, gci, bonus, transaction_fee, broker_split,
                franchise_split, referral_fee, primary_agent_id, primary_agent_name,
                primary_agent_pct, primary_agent_gci, secondary_agent_name,
                secondary_agent_pct, secondary_agent_gci, mortgage_company,
                title_company, location, net_income, taxes, net_after_taxes,
                year, month, created_at, updated_at
            ) VALUES (
                %s,%s,%s,%s,%s,%s, %s,%s,%s,%s, %s,%s,%s,%s,
                %s,%s,%s,%s,%s, %s,%s,%s,%s, %s,%s,%s, %s,%s,%s,
                %s,%s,%s,%s,%s, %s,%s,NOW(),NOW()
            )
        """, (
            status, sub_status, tx_type, address, cs(row[C_CLIENT],200), cs(row[C_SOURCE],100),
            cd(row[C_SIGNED]), cd(row[C_MLS_LIVE]), cd(row[C_EXP_DATE]), cd(row[C_UNDER_CONTR]),
            cd(row[C_PROJ_CLOSE]), close_dt, cm(row[C_LIST_PRICE]), cm(row[C_SALE_PRICE]),
            cp(row[C_COMM_PCT]), cm(row[C_GCI]) or 0.0, cm(row[C_BONUS]), cm(row[C_TX_FEE]), cm(row[C_BROKER_SPLIT]),
            cm(row[C_FRANCHISE]), cm(row[C_REFERRAL]), pri_id, pri_name,
            cp(row[C_PRI_PCT]), cm(row[C_PRIMARY_GCI]), cs(row[C_SEC_AGENT],100),
            cp(row[C_SEC_PCT]), cm(row[C_SEC_GCI]), cs(row[64],100),
            cs(row[65],100), cs(row[68],100), cm(row[72]), cm(row[73]), cm(row[74]),
            year, month
        ))
        imported += 1
        if imported % 50 == 0:
            conn.commit()
            print(f'  ... {imported} committed')
    except Exception as e:
        conn.rollback()
        if imported < 5: print(f'  ERR row {i}: {e}')

conn.commit()
print(f'  ✅ Transactions: imported={imported} skipped={skipped}')

# ── LEAD GEN ──────────────────────────────────────────────────────────
print('\n📞 Importing Lead Gen...')
ws2 = wb['Lead Gen']
rows2 = list(ws2.iter_rows(min_row=7, values_only=True))

lg_imported = lg_skipped = 0
for i, row in enumerate(rows2):
    if not row or len(row) < 3: lg_skipped += 1; continue
    log_date = cd(row[1])
    name     = cs(row[2], 100)
    if not log_date or not name or name == 'Name': lg_skipped += 1; continue

    agent_id = agents.get(name.strip().lower())
    if not agent_id:
        cur.execute("INSERT INTO agents (name, status, created_at) VALUES (%s,'Active',NOW()) RETURNING id", (name,))
        agent_id = cur.fetchone()[0]
        conn.commit()
        agents[name.strip().lower()] = agent_id

    def ci(v):
        try: return int(float(str(v))) if v else 0
        except: return 0
    def cf(v):
        try: return float(str(v)) if v else 0.0
        except: return 0.0

    try:
        cur.execute("""
            INSERT INTO lead_gen_log (
                agent_id, log_date, hours, dials, contacts, nurtures,
                listing_appts_set, listing_appts_held, listings_signed,
                buyer_appts_set, buyer_appts_held, buyers_signed,
                written_offers, showings, open_houses, lead_source, notes, created_at
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
        """, (
            agent_id, log_date, cf(row[3]), ci(row[4]), ci(row[5]), ci(row[6]),
            ci(row[7]), ci(row[8]), ci(row[9]),
            ci(row[10]), ci(row[11]), ci(row[12]),
            ci(row[13]), ci(row[14]), ci(row[15]),
            cs(row[18],50), cs(row[19])
        ))
        lg_imported += 1
        if lg_imported % 100 == 0:
            conn.commit()
            print(f'  ... {lg_imported} committed')
    except Exception as e:
        conn.rollback()
        if lg_imported < 5: print(f'  ERR row {i}: {e}')

conn.commit()
print(f'  ✅ Lead Gen: imported={lg_imported} skipped={lg_skipped}')

# ── Final counts ──────────────────────────────────────────────────────
cur.execute("SELECT COUNT(*) FROM transactions")
print(f'\n🎉 Done!')
print(f'   Transactions in DB : {cur.fetchone()[0]}')
cur.execute("SELECT COUNT(*) FROM lead_gen_log")
print(f'   Lead Gen logs in DB: {cur.fetchone()[0]}')
cur.execute("SELECT COUNT(*) FROM agents")
print(f'   Agents in DB       : {cur.fetchone()[0]}')
conn.close()
